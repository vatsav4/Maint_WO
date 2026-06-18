from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, io
from datetime import datetime

app = Flask(__name__)
DB = os.path.join(os.path.dirname(__file__), 'maintenance_db.xlsx')

MONTHS = ['APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB','MAR']
QTR_GROUPS = {'A':[0,3,6,9],'B':[1,4,7,10],'C':[2,5,8,11]}

# ── DB HELPERS ───────────────────────────────────────────────────────────────

def read_sheet(sheet_name):
    try:
        df = pd.read_excel(DB, sheet_name=sheet_name, dtype=str).fillna('')
        return df
    except Exception as e:
        return pd.DataFrame()

def get_next_id(df, id_col='ID'):
    if df.empty or id_col not in df.columns:
        return 1
    ids = pd.to_numeric(df[id_col], errors='coerce').dropna()
    return int(ids.max()) + 1 if len(ids) else 1

def append_row(sheet_name, row_dict):
    wb = load_workbook(DB)
    ws = wb[sheet_name]
    next_row = ws.max_row + 1
    for ci, val in enumerate(row_dict.values(), 1):
        c = ws.cell(row=next_row, column=ci, value=val)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        s = Side(style='thin')
        c.border = Border(left=s, right=s, top=s, bottom=s)
        bg = 'F2F2F2' if (next_row % 2 == 0) else 'FFFFFF'
        c.fill = PatternFill('solid', fgColor=bg)
    wb.save(DB)

def update_cell(sheet_name, id_val, col_name, new_val, id_col='ID'):
    update_row(sheet_name, id_val, {col_name: new_val}, id_col)

def update_row(sheet_name, id_val, fields: dict, id_col='ID'):
    wb = load_workbook(DB)
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    id_idx = headers.index(id_col) + 1
    col_map = {name: headers.index(name) + 1 for name in fields if name in headers}
    for row in ws.iter_rows(min_row=2):
        if str(row[id_idx-1].value) == str(id_val):
            for col_name, col_idx in col_map.items():
                row[col_idx-1].value = fields[col_name]
            break
    wb.save(DB)

def delete_row(sheet_name, id_val, id_col='ID'):
    """Soft delete by setting Active=NO"""
    update_cell(sheet_name, id_val, 'Active', 'NO', id_col)

def get_config():
    df = read_sheet('FY_Config')
    cfg = {}
    for _, r in df.iterrows():
        cfg[r.get('Setting','')] = r.get('Value','')
    return cfg

# ── SCHEDULE LOGIC ───────────────────────────────────────────────────────────

def get_tasks_for_month(month_idx):
    """Returns list of task dicts due in given month index (0=APR..11=MAR)"""
    tasks = []
    eq_df = read_sheet('Equipment')
    eq_df = eq_df[eq_df['Active']=='YES']

    for _, eq in eq_df.iterrows():
        name = eq.get('Equip_Name','')
        eqno = eq.get('Equip_No','')

        # Monthly PM
        if eq.get('PM_Monthly','')=='YES':
            tasks.append({
                'type':'PM-Monthly','ref':eq.get('Ref_Monthly',''),
                'equip_no':eqno,'equip_name':name,
                'frequency':'Monthly','task_name':f'Monthly PM – {name}',
                'maint_plan':eq.get('Maint_Plan_M',''),
                'task_list':eq.get('Task_List_M','')
            })

        # Quarterly PM
        if eq.get('PM_Quarterly','')=='YES':
            grp = eq.get('Quarterly_Group','A')
            if month_idx in QTR_GROUPS.get(grp,[]):
                tasks.append({
                    'type':'PM-Quarterly','ref':eq.get('Ref_Quarterly',''),
                    'equip_no':eqno,'equip_name':name,
                    'frequency':'Quarterly','task_name':f'Quarterly PM – {name}',
                    'maint_plan':'','task_list':''
                })

        # Half Yearly PM
        if eq.get('PM_Half_Yearly','')=='YES':
            hy_str = eq.get('HY_Months','')
            hy_months = [int(x) for x in hy_str.split(',') if x.strip().isdigit()]
            if month_idx in hy_months:
                tasks.append({
                    'type':'PM-HalfYearly','ref':eq.get('Ref_HalfYearly',''),
                    'equip_no':eqno,'equip_name':name,
                    'frequency':'Half Yearly','task_name':f'Half Yearly PM – {name}',
                    'maint_plan':'','task_list':''
                })

        # Yearly PM
        if eq.get('PM_Yearly','')=='YES':
            y_idx_str = eq.get('Yearly_Month_Index','')
            if y_idx_str.strip().isdigit() and int(y_idx_str) == month_idx:
                tasks.append({
                    'type':'PM-Yearly','ref':eq.get('Ref_Yearly',''),
                    'equip_no':eqno,'equip_name':name,
                    'frequency':'Yearly','task_name':f'Yearly PM – {name}',
                    'maint_plan':'','task_list':''
                })

    # TBM
    tbm_df = read_sheet('TBM')
    tbm_df = tbm_df[tbm_df['Active']=='YES']
    for _, t in tbm_df.iterrows():
        months_str = t.get('Due_Month_Indices','')
        months = [int(x) for x in months_str.split(',') if x.strip().isdigit()]
        if month_idx in months:
            tasks.append({
                'type':'TBM','ref':'','equip_no':'—',
                'equip_name':t.get('Equipment_Scope',''),
                'frequency':t.get('Frequency',''),
                'task_name':t.get('Task_Name',''),
                'maint_plan':'','task_list':''
            })

    # CBM
    cbm_df = read_sheet('CBM')
    cbm_df = cbm_df[cbm_df['Active']=='YES']
    for _, c in cbm_df.iterrows():
        months_str = c.get('Due_Month_Indices','')
        months = [int(x) for x in months_str.split(',') if x.strip().isdigit()]
        if month_idx in months:
            tasks.append({
                'type':'CBM','ref':'','equip_no':'—',
                'equip_name':c.get('Equipment_Scope',''),
                'frequency':c.get('Frequency',''),
                'task_name':c.get('Activity_Name',''),
                'maint_plan':'','task_list':''
            })

    return tasks

def get_log_status(fy, month_idx, task_type, task_name, equip_no):
    log = read_sheet('Activity_Log')
    if log.empty:
        return None, '', ''
    match = log[
        (log['FY']==str(fy)) &
        (log['Month_Index']==str(month_idx)) &
        (log['Task_Type']==task_type) &
        (log['Task_Name']==task_name) &
        (log['Equip_No']==str(equip_no))
    ]
    if match.empty:
        return None, '', ''
    row = match.iloc[-1]
    return row.get('Status','PENDING'), row.get('Remarks',''), row.get('Log_ID','')

def month_summary():
    summary = {}
    for mi, mon in enumerate(MONTHS):
        tasks = get_tasks_for_month(mi)
        by_type = {}
        for t in tasks:
            by_type[t['type']] = by_type.get(t['type'],0)+1
        summary[mon] = {'total':len(tasks),'by_type':by_type}
    return summary

# ── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    now = datetime.now()
    # Current FY month index: APR=0
    cal_month = now.month
    fy_month_idx = (cal_month - 4) % 12
    current_month = MONTHS[fy_month_idx]
    cfg = get_config()
    fy = cfg.get('Current_FY','2025-26')

    tasks = get_tasks_for_month(fy_month_idx)
    log = read_sheet('Activity_Log')

    done, pending, deferred = 0, 0, 0
    for t in tasks:
        status, _, _ = get_log_status(fy, fy_month_idx, t['type'], t['task_name'], t['equip_no'])
        if status == 'DONE': done += 1
        elif status == 'DEFERRED': deferred += 1
        else: pending += 1

    summary = month_summary()
    eq_count = len(read_sheet('Equipment').query("Active=='YES'")) if not read_sheet('Equipment').empty else 0
    tbm_count = len(read_sheet('TBM').query("Active=='YES'")) if not read_sheet('TBM').empty else 0
    cbm_count = len(read_sheet('CBM').query("Active=='YES'")) if not read_sheet('CBM').empty else 0

    return render_template('dashboard.html',
        fy=fy, current_month=current_month, fy_month_idx=fy_month_idx,
        total=len(tasks), done=done, pending=pending, deferred=deferred,
        summary=summary, months=MONTHS, eq_count=eq_count,
        tbm_count=tbm_count, cbm_count=cbm_count,
        completion_pct=round(done/len(tasks)*100) if tasks else 0)

@app.route('/calendar/<int:month_idx>')
def calendar_view(month_idx):
    cfg = get_config()
    fy = cfg.get('Current_FY','2025-26')
    tasks = get_tasks_for_month(month_idx)
    enriched = []
    for t in tasks:
        status, remarks, log_id = get_log_status(fy, month_idx, t['type'], t['task_name'], t['equip_no'])
        enriched.append({**t, 'status': status or 'PENDING', 'remarks': remarks, 'log_id': log_id})
    month_name = MONTHS[month_idx]

    done = sum(1 for t in enriched if t['status']=='DONE')
    pending = sum(1 for t in enriched if t['status']=='PENDING')
    deferred = sum(1 for t in enriched if t['status']=='DEFERRED')

    return render_template('calendar.html',
        fy=fy, month_name=month_name, month_idx=month_idx,
        tasks=enriched, months=MONTHS, done=done, pending=pending, deferred=deferred)

@app.route('/update_status', methods=['POST'])
def update_status():
    data = request.json
    cfg = get_config()
    fy = cfg.get('Current_FY','2025-26')
    month_idx = data['month_idx']
    log = read_sheet('Activity_Log')

    # Check if entry exists
    match = pd.DataFrame()
    if not log.empty:
        match = log[
            (log['FY']==str(fy)) &
            (log['Month_Index']==str(month_idx)) &
            (log['Task_Type']==data['task_type']) &
            (log['Task_Name']==data['task_name']) &
            (log['Equip_No']==str(data['equip_no']))
        ]

    if not match.empty:
        # Update existing
        log_id = match.iloc[-1]['Log_ID']
        update_cell('Activity_Log', log_id, 'Status', data['status'], 'Log_ID')
        update_cell('Activity_Log', log_id, 'Remarks', data.get('remarks',''), 'Log_ID')
        update_cell('Activity_Log', log_id, 'Updated_At', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Log_ID')
    else:
        # Insert new
        next_id = get_next_id(log, 'Log_ID')
        append_row('Activity_Log', {
            'Log_ID': next_id,
            'FY': fy,
            'Month_Index': month_idx,
            'Month_Name': MONTHS[month_idx],
            'Task_Type': data['task_type'],
            'Ref_No': data.get('ref',''),
            'Equip_No': data.get('equip_no',''),
            'Equip_Name': data.get('equip_name',''),
            'Task_Name': data['task_name'],
            'Frequency': data.get('frequency',''),
            'Status': data['status'],
            'Remarks': data.get('remarks',''),
            'Updated_At': datetime.now().strftime('%Y-%m-%d %H:%M')
        })
    return jsonify({'ok': True})

# ── EQUIPMENT CRUD ───────────────────────────────────────────────────────────

@app.route('/equipment')
def equipment():
    df = read_sheet('Equipment')
    active = df[df['Active']=='YES'] if not df.empty else df
    rows = active.to_dict('records')
    return render_template('equipment.html', rows=rows, months=MONTHS)

@app.route('/equipment/add', methods=['POST'])
def add_equipment():
    f = request.form
    df = read_sheet('Equipment')
    next_id = get_next_id(df)
    qgrp = f.get('quarterly_group','A') if f.get('pm_quarterly')=='YES' else ''
    hy_months = f.get('hy_months','') if f.get('pm_half_yearly')=='YES' else ''
    y_idx = f.get('yearly_month_idx','') if f.get('pm_yearly')=='YES' else ''
    append_row('Equipment', {
        'ID': next_id,
        'Equip_No': f.get('equip_no','').strip(),
        'Equip_Name': f.get('equip_name','').strip().upper(),
        'PM_Monthly': f.get('pm_monthly','NO'),
        'PM_Quarterly': f.get('pm_quarterly','NO'),
        'Quarterly_Group': qgrp,
        'PM_Half_Yearly': f.get('pm_half_yearly','NO'),
        'HY_Months': hy_months,
        'PM_Yearly': f.get('pm_yearly','NO'),
        'Yearly_Month_Index': y_idx,
        'Ref_Monthly': f.get('ref_monthly','').strip(),
        'Ref_Quarterly': f.get('ref_quarterly','').strip(),
        'Ref_HalfYearly': f.get('ref_halfyearly','').strip(),
        'Ref_Yearly': f.get('ref_yearly','').strip(),
        'Maint_Plan_M': f.get('maint_plan','').strip(),
        'Task_List_M': f.get('task_list','').strip(),
        'Active': 'YES'
    })
    return jsonify({'ok': True, 'id': next_id})

@app.route('/equipment/delete/<int:eq_id>', methods=['POST'])
def delete_equipment(eq_id):
    delete_row('Equipment', eq_id)
    return jsonify({'ok': True})

@app.route('/equipment/edit/<int:eq_id>', methods=['POST'])
def edit_equipment(eq_id):
    f = request.form
    fields = {
        'Equip_No': f.get('equip_no','').strip(),
        'Equip_Name': f.get('equip_name','').strip().upper(),
        'PM_Monthly': f.get('pm_monthly','NO'),
        'PM_Quarterly': f.get('pm_quarterly','NO'),
        'Quarterly_Group': f.get('quarterly_group','A') if f.get('pm_quarterly')=='YES' else '',
        'PM_Half_Yearly': f.get('pm_half_yearly','NO'),
        'HY_Months': f.get('hy_months','') if f.get('pm_half_yearly')=='YES' else '',
        'PM_Yearly': f.get('pm_yearly','NO'),
        'Yearly_Month_Index': f.get('yearly_month_idx','') if f.get('pm_yearly')=='YES' else '',
        'Ref_Monthly': f.get('ref_monthly',''),
        'Ref_Quarterly': f.get('ref_quarterly',''),
        'Ref_HalfYearly': f.get('ref_halfyearly',''),
        'Ref_Yearly': f.get('ref_yearly',''),
        'Maint_Plan_M': f.get('maint_plan',''),
        'Task_List_M': f.get('task_list',''),
    }
    update_row('Equipment', eq_id, fields)
    return jsonify({'ok': True})

# ── TBM CRUD ─────────────────────────────────────────────────────────────────

@app.route('/tbm')
def tbm_view():
    df = read_sheet('TBM')
    active = df[df['Active']=='YES'] if not df.empty else df
    rows = active.to_dict('records')
    return render_template('tbm_cbm.html', rows=rows, sheet='TBM',
                           title='Time Based Maintenance', months=MONTHS)

@app.route('/tbm/add', methods=['POST'])
def add_tbm():
    f = request.form
    df = read_sheet('TBM')
    next_id = get_next_id(df)
    append_row('TBM', {
        'ID': next_id,
        'Task_Name': f.get('task_name','').strip(),
        'Frequency': f.get('frequency','').strip(),
        'Due_Month_Indices': f.get('due_months','').strip(),
        'Equipment_Scope': f.get('equipment_scope','').strip(),
        'Active': 'YES'
    })
    return jsonify({'ok': True})

@app.route('/tbm/delete/<int:tid>', methods=['POST'])
def delete_tbm(tid):
    delete_row('TBM', tid)
    return jsonify({'ok': True})

@app.route('/tbm/edit/<int:tid>', methods=['POST'])
def edit_tbm(tid):
    f = request.form
    for col, key in [('Task_Name','task_name'),('Frequency','frequency'),
                     ('Due_Month_Indices','due_months'),('Equipment_Scope','equipment_scope')]:
        update_cell('TBM', tid, col, f.get(key,'').strip())
    return jsonify({'ok': True})

# ── CBM CRUD ─────────────────────────────────────────────────────────────────

@app.route('/cbm')
def cbm_view():
    df = read_sheet('CBM')
    active = df[df['Active']=='YES'] if not df.empty else df
    rows = active.to_dict('records')
    return render_template('tbm_cbm.html', rows=rows, sheet='CBM',
                           title='Condition Based Maintenance', months=MONTHS)

@app.route('/cbm/add', methods=['POST'])
def add_cbm():
    f = request.form
    df = read_sheet('CBM')
    next_id = get_next_id(df)
    append_row('CBM', {
        'ID': next_id,
        'Activity_Name': f.get('task_name','').strip(),
        'Frequency': f.get('frequency','').strip(),
        'Due_Month_Indices': f.get('due_months','').strip(),
        'Equipment_Scope': f.get('equipment_scope','').strip(),
        'Active': 'YES'
    })
    return jsonify({'ok': True})

@app.route('/cbm/delete/<int:cid>', methods=['POST'])
def delete_cbm(cid):
    delete_row('CBM', cid)
    return jsonify({'ok': True})

@app.route('/cbm/edit/<int:cid>', methods=['POST'])
def edit_cbm(cid):
    f = request.form
    name_col = 'Activity_Name'
    for col, key in [(name_col,'task_name'),('Frequency','frequency'),
                     ('Due_Month_Indices','due_months'),('Equipment_Scope','equipment_scope')]:
        update_cell('CBM', cid, col, f.get(key,'').strip())
    return jsonify({'ok': True})

# ── EXPORT ───────────────────────────────────────────────────────────────────

@app.route('/export/<string:scope>')
def export_excel(scope):
    cfg = get_config()
    fy = cfg.get('Current_FY','2025-26')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Maintenance Calendar'
    ws.sheet_view.showGridLines = False

    def hdr_cell(row, col, val, bg='1F3864', fg='FFFFFF', bold=True, w=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(bold=bold, color=fg, name='Arial', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        s = Side(style='thin')
        c.border = Border(left=s,right=s,top=s,bottom=s)
        if w: ws.column_dimensions[get_column_letter(col)].width = w

    TYPE_BG = {
        'PM-Monthly':'BDD7EE','PM-Quarterly':'E2EFDA',
        'PM-HalfYearly':'E2D9F3','PM-Yearly':'FCE4D6',
        'TBM':'E0F7FA','CBM':'FFE0E0'
    }

    log = read_sheet('Activity_Log')

    if scope == 'year':
        # Full year export with month columns
        ws.merge_cells('A1:R1')
        t = ws['A1']
        t.value = f'MHCV MAINTENANCE CALENDAR  —  FY {fy}'
        t.fill = PatternFill('solid',fgColor='1F3864')
        t.font = Font(bold=True,color='FFFFFF',name='Arial',size=13)
        t.alignment = Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[1].height=28

        headers = ['#','Task Type','Ref No','Equip No','Equipment / Task','Frequency',
                   'APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB','MAR']
        widths =  [5,  14,      12,    10,   32,              13,
                   6,  6,  6,  6,  6,  6,  6,  6,  6,  6,  6,  6]
        for i,(h,w) in enumerate(zip(headers,widths),1):
            hdr_cell(2,i,h,w=w)
        ws.row_dimensions[2].height=22

        r=3; sr=0
        for mi in range(12):
            tasks = get_tasks_for_month(mi)
            for t in tasks:
                # avoid duplicates in annual view (monthly PM shown once)
                if t['type']=='PM-Monthly' and mi>0: continue
                sr+=1
                due_months=[mi]
                if t['type']=='PM-Monthly': due_months=list(range(12))
                tbg=TYPE_BG.get(t['type'],'F2F2F2')
                bg='F2F2F2' if sr%2==0 else 'FFFFFF'
                vals=[sr,t['type'],t['ref'],t['equip_no'],t['task_name'],t['frequency']]
                for ci,v in enumerate(vals,1):
                    c=ws.cell(row=r,column=ci,value=v)
                    c.fill=PatternFill('solid',fgColor=tbg if ci==2 else bg)
                    c.font=Font(name='Arial',size=9)
                    c.alignment=Alignment(horizontal='left' if ci==5 else 'center',vertical='center')
                    s=Side(style='thin'); c.border=Border(left=s,right=s,top=s,bottom=s)
                for mi2 in range(12):
                    col=7+mi2
                    in_due = mi2 in due_months
                    cell=ws.cell(row=r,column=col,value='✔' if in_due else '')
                    cell.fill=PatternFill('solid',fgColor=tbg if in_due else bg)
                    cell.font=Font(name='Arial',size=10 if in_due else 9,
                                   bold=in_due,color='1F3864' if in_due else '999999')
                    cell.alignment=Alignment(horizontal='center',vertical='center')
                    s=Side(style='thin'); cell.border=Border(left=s,right=s,top=s,bottom=s)
                ws.row_dimensions[r].height=15
                r+=1
    else:
        # Monthly export
        month_idx = int(scope)
        month_name = MONTHS[month_idx]
        tasks = get_tasks_for_month(month_idx)

        ws.merge_cells('A1:K1')
        t=ws['A1']
        t.value=f'MHCV MAINTENANCE — {month_name}  |  FY {fy}'
        t.fill=PatternFill('solid',fgColor='1F3864')
        t.font=Font(bold=True,color='FFFFFF',name='Arial',size=13)
        t.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[1].height=28

        headers=['#','Task Type','Ref No','Equip No','Equipment / Task','Frequency','Maint Plan','Task List','Status','Remarks']
        widths= [5,  14,      12,    10,   32,              13,          14,          12,          12,       30]
        for i,(h,w) in enumerate(zip(headers,widths),1):
            hdr_cell(2,i,h,w=w)
        ws.row_dimensions[2].height=22

        r=3
        for sr,t in enumerate(tasks,1):
            status,remarks,_ = get_log_status(fy,month_idx,t['type'],t['task_name'],t['equip_no'])
            status=status or 'PENDING'
            tbg=TYPE_BG.get(t['type'],'F2F2F2')
            bg='F2F2F2' if sr%2==0 else 'FFFFFF'
            sbg={'DONE':'E2EFDA','DEFERRED':'FFE0E0'}.get(status,'FFF2CC')
            vals=[sr,t['type'],t['ref'],t['equip_no'],t['task_name'],t['frequency'],
                  t['maint_plan'],t['task_list'],status,remarks]
            for ci,v in enumerate(vals,1):
                c=ws.cell(row=r,column=ci,value=v)
                c.fill=PatternFill('solid',fgColor=tbg if ci==2 else (sbg if ci==9 else bg))
                c.font=Font(name='Arial',size=9,bold=(ci==9))
                c.alignment=Alignment(horizontal='left' if ci in [5,10] else 'center',vertical='center')
                s=Side(style='thin'); c.border=Border(left=s,right=s,top=s,bottom=s)
            ws.row_dimensions[r].height=15
            r+=1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'MHCV_Maintenance_{"Full_Year" if scope=="year" else MONTHS[int(scope)]}_{fy}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/equipment')
def api_equipment():
    df = read_sheet('Equipment')
    active = df[df['Active']=='YES'] if not df.empty else df
    return jsonify(active.to_dict('records'))

@app.route('/api/summary')
def api_summary():
    return jsonify(month_summary())

if __name__ == '__main__':
    if not os.path.exists(DB):
        print('DB not found. Run init_db.py first.')
    app.run(debug=True, host='0.0.0.0', port=5000)
