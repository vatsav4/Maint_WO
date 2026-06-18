"""
Run once to create maintenance_db.xlsx with seed data.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

DB_PATH = os.path.join(os.path.dirname(__file__), 'maintenance_db.xlsx')

# ── SEED DATA ──────────────────────────────────────────────────────────────

EQUIPMENT = [
    # (equip_no, equip_name, pm_monthly, pm_quarterly, pm_half_yearly, pm_yearly, ref_monthly, ref_quarterly, ref_half_yearly, ref_yearly, maint_plan_m, task_list_m)
    ('7001509','CONVEYOR',True,True,False,False,'MHCV/M/PM 01','MHCV/Q/PM 01','','','3000003047','18001'),
    ('7001510','SINGLE GIRDER CRANE 01',True,True,False,True,'MHCV/M/PM 02','MHCV/Q/PM 02','','MHCV/Y/PM 01','3000003031','17988'),
    ('7001511','SINGLE GIRDER CRANE 02',True,True,False,True,'MHCV/M/PM 03','MHCV/Q/PM 03','','MHCV/Y/PM 02','3000003032','17988'),
    ('7001512','SINGLE GIRDER CRANE 03',True,True,False,True,'MHCV/M/PM 04','MHCV/Q/PM 04','','MHCV/Y/PM 03','3000003041','17988'),
    ('7001513','SINGLE GIRDER CRANE 04',True,True,False,True,'MHCV/M/PM 05','MHCV/Q/PM 05','','MHCV/Y/PM 04','3000003042','17988'),
    ('7001514','SINGLE GIRDER CRANE 05',True,True,False,True,'MHCV/M/PM 06','MHCV/Q/PM 06','','MHCV/Y/PM 05','3000003043','17988'),
    ('7001515','SINGLE GIRDER CRANE 06',True,True,False,True,'MHCV/M/PM 07','MHCV/Q/PM 07','','MHCV/Y/PM 06','3000003044','17988'),
    ('7001516','SINGLE GIRDER CRANE 07',True,True,False,True,'MHCV/M/PM 08','MHCV/Q/PM 08','','MHCV/Y/PM 07','3000003045','17988'),
    ('7001517','SINGLE GIRDER CRANE 08',True,True,False,True,'MHCV/M/PM 09','MHCV/Q/PM 09','','MHCV/Y/PM 08','3000003048','18010'),
    ('7001518','SINGLE GIRDER CRANE 09',True,True,False,True,'MHCV/M/PM 10','MHCV/Q/PM 10','','MHCV/Y/PM 09','3000003049','18010'),
    ('7001519','SINGLE GIRDER CRANE 10',True,True,False,True,'MHCV/M/PM 11','MHCV/Q/PM 11','','MHCV/Y/PM 10','3000003050','18010'),
    ('7001507','DOUBLE GIRDER CRANE 01',True,True,False,True,'MHCV/M/PM 12','MHCV/Q/PM 12','','MHCV/Y/PM 11','3000003046','17994'),
    ('7001508','DOUBLE GIRDER CRANE 02',True,True,False,True,'MHCV/M/PM 13','MHCV/Q/PM 13','','MHCV/Y/PM 12','3000003051','17994'),
    ('7001520','DOUBLE GIRDER CRANE 03',True,True,False,True,'MHCV/M/PM 14','MHCV/Q/PM 14','','MHCV/Y/PM 13','3000003052','17994'),
    ('7001535','GOLIATH CRANE 01',True,True,False,True,'MHCV/M/PM 15','MHCV/Q/PM 15','','MHCV/Y/PM 14','3000003071','18014'),
    ('7001572','GOLIATH CRANE 02',True,True,False,True,'MHCV/M/PM 16','MHCV/Q/PM 16','','MHCV/Y/PM 15','3000003072','18015'),
    ('7001570','HOIST 01',True,True,False,False,'MHCV/M/PM 17','MHCV/Q/PM 17','','','3000003061','18007'),
    ('7001523/24','INVERSION',True,False,False,True,'MHCV/M/PM 18','','','MHCV/Y/PM 27','3000003087','18016'),
    ('7001521','U-BOLT MACHINE LH',True,False,True,True,'MHCV/M/PM 19','','MHCV/HY/PM 02','MHCV/Y/PM 18','3000003053','18002'),
    ('7001522','U-BOLT MACHINE RH',True,False,True,True,'MHCV/M/PM 20','','MHCV/HY/PM 01','MHCV/Y/PM 19','3000003054','18002'),
    ('7001525','MANIPULATOR LH',True,True,True,True,'MHCV/M/PM 21','MHCV/Q/PM 20','MHCV/HY/PM 03','MHCV/Y/PM 21','3000003055','18004'),
    ('7001526','MANIPULATOR RH',True,True,True,True,'MHCV/M/PM 22','MHCV/Q/PM 21','MHCV/HY/PM 04','MHCV/Y/PM 20','3000003056','18004'),
    ('7001527','WHEEL NUT RUNNER LH',True,False,True,True,'MHCV/M/PM 23','','MHCV/HY/PM 05','MHCV/Y/PM 16','3000003057','18003'),
    ('7001528','WHEEL NUT RUNNER RH',True,False,True,True,'MHCV/M/PM 24','','MHCV/HY/PM 06','MHCV/Y/PM 17','3000003058','18003'),
    ('7001531','GREASE DISPENSING MC',True,True,False,False,'MHCV/M/PM 25','MHCV/Q/PM 22','','','3000003081','18017'),
    ('7001533','CHASSIS NO. PUNCHING MC',True,True,False,True,'MHCV/M/PM 26','MHCV/Q/PM 23','','MHCV/Y/PM 23','3000003082','18018'),
    ('7001532','PLATE NO. PUNCHING MC',True,True,False,True,'MHCV/M/PM 27','MHCV/Q/PM 24','','MHCV/Y/PM 22','3000003083','18019'),
    ('7001536','SMOKE TESTING MC',True,False,True,False,'MHCV/M/PM 28','','MHCV/HY/PM 09','','3000003084','18020'),
    ('7001529','UREA & COOLANT FILLING MC',True,True,False,False,'MHCV/M/PM 29','MHCV/Q/PM 25','','','3000003085','18021'),
    ('7001530','CLUTCH & PAS FILLING MC',True,True,False,False,'MHCV/M/PM 30','MHCV/Q/PM 26','','','3000003086','18022'),
    ('7001537','WHEEL ALIGNMENT MC',True,False,False,True,'MHCV/M/PM 31','','','MHCV/Y/PM 28','3000003090','18023'),
    ('7001571','HOIST 02',True,True,False,False,'MHCV/M/PM 32','MHCV/Q/PM 18','','','3000003062','18007'),
    ('7001574','ZIB CRANE 01',True,True,False,False,'MHCV/M/PM 33','MHCV/Q/PM 28','','','3000003073','18005'),
    ('7001575','ZIB CRANE 02',True,True,False,False,'MHCV/M/PM 34','MHCV/Q/PM 29','','','3000003074','18005'),
    ('7001576','ZIB CRANE 03',True,True,False,False,'MHCV/M/PM 35','MHCV/Q/PM 30','','','3000003075','18005'),
    ('7001578','ZIB CRANE 04',True,True,False,False,'MHCV/M/PM 36','MHCV/Q/PM 31','','','3000003091','18025'),
    ('7001569','EOL PIT 1',True,False,False,False,'MHCV/M/PM 37','','','','3000003076','18005'),
    ('7001599','Y1 MACHINE',True,False,False,True,'MHCV/M/PM 38','','','MHCV/Y/PM 29','3000003269','18341'),
    ('7001600','TYRE ASSEMBLY MC',True,False,False,True,'MHCV/M/PM 39','','','MHCV/Y/PM 25','3000003270','18342'),
    ('7001601','TYRE INFLATOR',True,False,False,True,'MHCV/M/PM 40','','','MHCV/Y/PM 26','3000003271','18343'),
    ('7001534','TOOL TROLLEY',True,False,False,False,'MHCV/M/PM 41','','','','3000003261','18335'),
    ('7001604','INSULATION CHECKING ST',True,False,False,False,'MHCV/M/PM 42','','','','3000003263','18337'),
    ('7001538','SUB ASSY 1',True,False,False,False,'MHCV/M/PM 43','','','','3000003264','18338'),
    ('7001539','SUB ASSY 2',True,False,False,False,'MHCV/M/PM 44','','','','3000003265','18338'),
    ('7001603','ECOS STATION',True,False,False,True,'MHCV/M/PM 45','','','MHCV/Y/PM 30','3000003262','18336'),
    ('7001740','RBT',True,False,False,False,'MHCV/M/PM 46','','','','','18965'),
    ('7001746','SG 11',True,False,False,False,'MHCV/M/PM 47','','','','3000003751','18968'),
    ('—','TYRE SHED WHEEL BALANCER',True,False,False,False,'MHCV/M/PM 48','','','','3000003752',''),
    ('—','EMERGENCY LAMP TEST',True,False,False,False,'MHCV/M/PM 49','','','','3000003753','18970'),
    ('7001607','DOCK LEVELLER',False,True,False,False,'','MHCV/Q/PM 27','','','3000003179','18047'),
    ('7001577','BEARING PRESSING MC',False,True,False,True,'','MHCV/Q/PM 32','','MHCV/Y/PM 31','3000003111','18110'),
    ('7001568','AXLE ALIGNMENT MC',False,True,False,True,'','MHCV/Q/PM 33','','MHCV/Y/PM 24','3000003089','18024'),
    ('7001602','ECU FLASHING',False,True,False,True,'','MHCV/Q/PM 34','','MHCV/Y/PM 30','3000003266','18339'),
    ('7001605','AC FILLING 1',False,True,False,False,'','MHCV/Q/PM 35','','','3000003267','18340'),
    ('7001606','AC FILLING 2',False,True,False,False,'','MHCV/Q/PM 36','','','3000003268','18340'),
    ('7001608','POWER PANEL 1',False,False,True,False,'','','MHCV/HY/PM 07','','3000003148','18112'),
    ('7001609','POWER PANEL 2',False,False,True,False,'','','MHCV/HY/PM 08','','3000003149',''),
]

# Quarterly group mapping (A=APR/JUL/OCT/JAN, B=MAY/AUG/NOV/FEB, C=JUN/SEP/DEC/MAR)
QTR_GROUP = {
    'MHCV/Q/PM 01':'A','MHCV/Q/PM 02':'A','MHCV/Q/PM 03':'A',
    'MHCV/Q/PM 04':'B','MHCV/Q/PM 05':'B','MHCV/Q/PM 06':'C',
    'MHCV/Q/PM 07':'C','MHCV/Q/PM 08':'A','MHCV/Q/PM 09':'A',
    'MHCV/Q/PM 10':'B','MHCV/Q/PM 11':'B','MHCV/Q/PM 12':'C',
    'MHCV/Q/PM 13':'A','MHCV/Q/PM 14':'B','MHCV/Q/PM 15':'B',
    'MHCV/Q/PM 16':'B','MHCV/Q/PM 17':'C','MHCV/Q/PM 18':'A',
    'MHCV/Q/PM 19':'C','MHCV/Q/PM 20':'C','MHCV/Q/PM 21':'C',
    'MHCV/Q/PM 22':'C','MHCV/Q/PM 23':'C','MHCV/Q/PM 24':'C',
    'MHCV/Q/PM 25':'C','MHCV/Q/PM 26':'C','MHCV/Q/PM 27':'C',
    'MHCV/Q/PM 28':'C','MHCV/Q/PM 29':'C','MHCV/Q/PM 30':'C',
    'MHCV/Q/PM 31':'C','MHCV/Q/PM 32':'A','MHCV/Q/PM 33':'A',
    'MHCV/Q/PM 34':'A','MHCV/Q/PM 35':'A','MHCV/Q/PM 36':'A',
}

# Half yearly: months (0=APR..11=MAR)
HY_MONTHS = {
    'MHCV/HY/PM 01':'2,8','MHCV/HY/PM 02':'2,8','MHCV/HY/PM 03':'2,8',
    'MHCV/HY/PM 04':'2,8','MHCV/HY/PM 05':'2,8','MHCV/HY/PM 06':'2,8',
    'MHCV/HY/PM 07':'3,9','MHCV/HY/PM 08':'3,9','MHCV/HY/PM 09':'0,6',
}

# Yearly: month index (0=APR)
Y_MONTH = {
    'MHCV/Y/PM 01':0,'MHCV/Y/PM 02':1,'MHCV/Y/PM 03':2,'MHCV/Y/PM 04':3,
    'MHCV/Y/PM 05':4,'MHCV/Y/PM 06':5,'MHCV/Y/PM 07':6,'MHCV/Y/PM 08':7,
    'MHCV/Y/PM 09':8,'MHCV/Y/PM 10':9,'MHCV/Y/PM 11':10,'MHCV/Y/PM 12':11,
    'MHCV/Y/PM 13':2,'MHCV/Y/PM 14':2,'MHCV/Y/PM 15':2,'MHCV/Y/PM 16':4,
    'MHCV/Y/PM 17':4,'MHCV/Y/PM 18':3,'MHCV/Y/PM 19':3,'MHCV/Y/PM 20':4,
    'MHCV/Y/PM 21':4,'MHCV/Y/PM 22':5,'MHCV/Y/PM 23':6,'MHCV/Y/PM 24':7,
    'MHCV/Y/PM 25':0,'MHCV/Y/PM 26':0,'MHCV/Y/PM 27':0,'MHCV/Y/PM 28':1,
    'MHCV/Y/PM 29':1,'MHCV/Y/PM 30':2,'MHCV/Y/PM 31':3,
}

TBM = [
    ('Manipulator Brake Pads Replacement','Quarterly','0,3,6,9','Manipulator LH/RH'),
    ('Filling Machine Gun/Adapter Seal Changing','Half Yearly','4,10','Urea/Coolant & Clutch Filling MC'),
    ('Coolant Tank Cleaning Activity','Yearly','2','Coolant Filling MC'),
    ('Stylus, Spring & O-ring Replacement – No. Punching MC','Yearly','7','Chassis/Plate No. Punching MC'),
    ('Vacuum Pump Overhauling','Yearly','9','Smoke Testing MC'),
    ('Grease Dispensing MC Gun Nozzle Replacement','Yearly','8','Grease Dispensing MC'),
    ('Tyre Inflator Nozzle Replacement','Yearly','1','Tyre Inflator'),
    ('NRV Replacement – Clutch & Brake Oil Filling','2 Yearly','2','Clutch & PAS Filling MC'),
    ('Conveyor Main Motor Gearbox Oil Replacement (SERVO-220)','2 Yearly','5','Conveyor'),
    ('Grease Dispensing MC High Pressure Hose Replacement','2 Yearly','8','Grease Dispensing MC'),
    ('Tyre Inflator High Pressure Hose Replacement','2 Yearly','1','Tyre Inflator'),
    ('Number Punching LM Guide Rail Changing','3 Yearly','2','Chassis/Plate No. Punching MC'),
    ('Conveyor Power Contactor Replacement','5 Yearly','11','Conveyor'),
]

CBM = [
    ('Motor Current Monitoring','Half Yearly','0,6','All Drive Motors'),
    ('All Hoists Chain Condition Check','Quarterly','0,3,6,9','Hoist 01, Hoist 02'),
    ('All Cranes Rope Condition Check','Quarterly','0,3,6,9','All Cranes'),
    ('Conveyor Trolley & Tackle Pads Replacement','Quarterly','0,3,6,9','Conveyor'),
    ('Air Gap / Brake Gap Setting – All Hoists','Yearly','0','Hoist 01, Hoist 02'),
    ('Lux Level Monitoring','Quarterly','0,3,6,9','Shop Floor'),
]

# ── BUILD EXCEL ──────────────────────────────────────────────────────────────

def hdr(ws, row, col, val, bg='1F3864', fg='FFFFFF', bold=True, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(bold=bold, color=fg, name='Arial', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    s = Side(style='thin')
    c.border = Border(left=s, right=s, top=s, bottom=s)
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w

wb = Workbook()

# ── Sheet 1: Equipment ────────────────────────────────────────────────────
ws_eq = wb.active
ws_eq.title = 'Equipment'
cols_eq = ['ID','Equip_No','Equip_Name','PM_Monthly','PM_Quarterly','Quarterly_Group',
           'PM_Half_Yearly','HY_Months','PM_Yearly','Yearly_Month_Index',
           'Ref_Monthly','Ref_Quarterly','Ref_HalfYearly','Ref_Yearly',
           'Maint_Plan_M','Task_List_M','Active']
widths = [5,12,30,11,13,14,15,10,10,18,14,14,14,12,14,12,7]
for i,(h,w) in enumerate(zip(cols_eq,widths),1):
    hdr(ws_eq,1,i,h,w=w)
ws_eq.row_dimensions[1].height = 22

for idx, row in enumerate(EQUIPMENT, 1):
    eqno, name, pm_m, pm_q, pm_hy, pm_y, ref_m, ref_q, ref_hy, ref_y, mp, tl = row
    qgrp = QTR_GROUP.get(ref_q,'A') if pm_q and ref_q else ''
    hy_months = HY_MONTHS.get(ref_hy,'') if pm_hy and ref_hy else ''
    y_idx = Y_MONTH.get(ref_y,'') if pm_y and ref_y else ''
    data = [idx, eqno, name,
            'YES' if pm_m else 'NO',
            'YES' if pm_q else 'NO', qgrp,
            'YES' if pm_hy else 'NO', hy_months,
            'YES' if pm_y else 'NO', y_idx,
            ref_m, ref_q, ref_hy, ref_y, mp, tl, 'YES']
    for ci, v in enumerate(data, 1):
        c = ws_eq.cell(row=idx+1, column=ci, value=v)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='center' if ci not in [3] else 'left',
                                vertical='center')
        s = Side(style='thin')
        c.border = Border(left=s,right=s,top=s,bottom=s)
        bg = 'F2F2F2' if idx%2==0 else 'FFFFFF'
        c.fill = PatternFill('solid', fgColor=bg)
    ws_eq.row_dimensions[idx+1].height = 15

# ── Sheet 2: TBM ─────────────────────────────────────────────────────────
ws_tbm = wb.create_sheet('TBM')
cols_tbm = ['ID','Task_Name','Frequency','Due_Month_Indices','Equipment_Scope','Active']
widths_tbm = [5,50,12,18,35,7]
for i,(h,w) in enumerate(zip(cols_tbm,widths_tbm),1):
    hdr(ws_tbm,1,i,h,w=w)
ws_tbm.row_dimensions[1].height = 22

for idx,(name,freq,months,scope) in enumerate(TBM,1):
    data=[idx,name,freq,months,scope,'YES']
    for ci,v in enumerate(data,1):
        c=ws_tbm.cell(row=idx+1,column=ci,value=v)
        c.font=Font(name='Arial',size=9)
        c.alignment=Alignment(horizontal='left' if ci in [2,5] else 'center',vertical='center')
        s=Side(style='thin'); c.border=Border(left=s,right=s,top=s,bottom=s)
        c.fill=PatternFill('solid',fgColor='F2F2F2' if idx%2==0 else 'FFFFFF')
    ws_tbm.row_dimensions[idx+1].height=15

# ── Sheet 3: CBM ─────────────────────────────────────────────────────────
ws_cbm = wb.create_sheet('CBM')
cols_cbm = ['ID','Activity_Name','Frequency','Due_Month_Indices','Equipment_Scope','Active']
widths_cbm = [5,50,12,18,35,7]
for i,(h,w) in enumerate(zip(cols_cbm,widths_cbm),1):
    hdr(ws_cbm,1,i,h,w=w)
ws_cbm.row_dimensions[1].height=22

for idx,(name,freq,months,scope) in enumerate(CBM,1):
    data=[idx,name,freq,months,scope,'YES']
    for ci,v in enumerate(data,1):
        c=ws_cbm.cell(row=idx+1,column=ci,value=v)
        c.font=Font(name='Arial',size=9)
        c.alignment=Alignment(horizontal='left' if ci in [2,5] else 'center',vertical='center')
        s=Side(style='thin'); c.border=Border(left=s,right=s,top=s,bottom=s)
        c.fill=PatternFill('solid',fgColor='F2F2F2' if idx%2==0 else 'FFFFFF')
    ws_cbm.row_dimensions[idx+1].height=15

# ── Sheet 4: Activity_Log ────────────────────────────────────────────────
ws_log = wb.create_sheet('Activity_Log')
cols_log=['Log_ID','FY','Month_Index','Month_Name','Task_Type','Ref_No','Equip_No',
          'Equip_Name','Task_Name','Frequency','Status','Remarks','Updated_At']
widths_log=[8,8,10,10,12,14,12,28,40,12,12,30,20]
for i,(h,w) in enumerate(zip(cols_log,widths_log),1):
    hdr(ws_log,1,i,h,w=w)
ws_log.row_dimensions[1].height=22

# ── Sheet 5: FY_Config ───────────────────────────────────────────────────
ws_cfg = wb.create_sheet('FY_Config')
ws_cfg['A1']='Setting'; ws_cfg['B1']='Value'
for c in [ws_cfg['A1'],ws_cfg['B1']]:
    c.fill=PatternFill('solid',fgColor='1F3864')
    c.font=Font(bold=True,color='FFFFFF',name='Arial',size=9)
ws_cfg['A2']='Current_FY'; ws_cfg['B2']='2025-26'
ws_cfg['A3']='FY_Start_Month'; ws_cfg['B3']='APR'
ws_cfg['A4']='App_Version'; ws_cfg['B4']='1.0'

wb.save(DB_PATH)
print(f'Database created: {DB_PATH}')
print(f'Equipment: {len(EQUIPMENT)} | TBM: {len(TBM)} | CBM: {len(CBM)}')
